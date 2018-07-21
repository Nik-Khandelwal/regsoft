from __future__ import unicode_literals
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from main.models import Group,Regplayer,Enteredplayer,Sport,Team
from django.contrib.auth import logout
from django.contrib.auth.models import User
import json
from django.contrib.auth import get_user_model
import string
import random
from django.core import serializers
import pusher
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.shortcuts import render
from django.views.decorators.cache import cache_page
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import StringIO
#import cStringIO as StringIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context
from django.http import HttpResponse
from cgi import escape
#from weasyprint import HTML
from random import choice
from string import ascii_uppercase
import re
import openpyxl
from openpyxl.utils import get_column_letter
import csv
from django.utils.encoding import smart_str
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404,HttpResponseRedirect, JsonResponse
from main.models import CustomUser, Team
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.views.generic import View, ListView, FormView
from django.views import generic
from register.forms import LoginForm, TeamForm, PlayerForm
from django.db import IntegrityError
from django.contrib.auth.models import User
import json
from django.core import serializers, mail
from django.contrib.auth.decorators import login_required
from django.dispatch import receiver

from random import choice
from string import ascii_uppercase

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from register.tokens import account_activation_token
from django.core.mail import EmailMessage
from django.core import serializers
from main.models import Group,Regplayer,Enteredplayer,Sport,Money,Billcontrols,Controls_user
from datetime import datetime

from django.contrib.auth import get_user_model
User=get_user_model()
 
 

CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)

pusher_client = pusher.Pusher(
  app_id='499153',
  key='9b825df805e0b694cccc',
  secret='f2bbd60c69e36c90a572',
  cluster='ap2',
  ssl=True
)

def is_controls_admin(user):
	if user:
		if Controls_user.objects.get(pk=1).user == user:
			return True
	return False


@cache_page(CACHE_TTL)
@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def main(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	mon = Money()
	mon.save()
	return render(request,'controls/index.html')


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def details(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data=[]
	for gr in Group.objects.all():
		b=[]
		pl = Enteredplayer.objects.filter(group=gr).filter(controls_passed=False)
		a=[]
		for p in pl:
			a.append(Regplayer.objects.get(pk=p.regplayer_id))
			p.controls_displayed = True
			p.save()
		for t in a:
			s=[]
			s.append(t.city)
			s.append(t.name.name)
			s.append(t.email_id)
			s.append(t.gender)
			s.append(t.unbilled_amt)
			s.append(t.college)
			s.append(t.mobile_no)
			s.append(t.entered)
			s.append(t.sport)
			s.append(t.pk)
			b.append(s)
		if b:
			data.append({"participants":b,"groupid":gr.group_code})
	return HttpResponse(json.dumps(data), content_type='application/json')


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def create_bill(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	bil = Billcontrols()
	bil.save()
	dat = {"success":1}
	return HttpResponse(json.dumps(dat), content_type='application/json')



@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def generate_bill(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		sums = 0
		print(data['data']['id_arr'])
		for i in data['data']['id_arr']:
			rp = Regplayer.objects.get(pk=int(i))
			pl = Enteredplayer.objects.get(regplayer = rp)
			pl.controls_passed = True
			pl.billcontrols = Billcontrols.objects.filter(unbilled_amt=0).first()
			pl.save()
		bil = Billcontrols.objects.filter(unbilled_amt=0).first()
		bil.unbilled_amt = data['data']['net_amt']
		bil.amt_received = int(data['data']['deno_2000'])*2000 + int(data['data']['deno_500'])*500 + int(data['data']['deno_200'])*200 + int(data['data']['deno_100'])*100 + int(data['data']['deno_50'])*50
		bil.save()
		money = Money.objects.get(pk=1)
		print(data)
		money.twothousand += int(data['data']['deno_2000'])
		money.fivehundred += int(data['data']['deno_500'])
		money.twohundred += int(data['data']['deno_200'])
		money.hundred += int(data['data']['deno_100'])
		money.fifty += int(data['data']['deno_50'])
		money.save()

		dat = {"success":1, "bills_pk":bil.pk}
		print(pl.controls_passed)



		data_recnacc=[]
		for gr in Group.objects.all():
			b=[]
			pl = Enteredplayer.objects.filter(controls_passed=True).filter(group=gr).filter(recnacc_displayed=False)
			a=[]
			for p in pl:
				a.append(Regplayer.objects.get(pk=p.regplayer_id))
				p.recnacc_displayed = True
				p.save()
			for t in a:
				b.append({"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
			if b:
				data_recnacc.append({"participants":b,"groupid":gr.group_code})
				
		print("pusher start")
		pusher_client.trigger('my-channel', 'my-event', data_recnacc)
		print("pusher")

		return HttpResponse(json.dumps(dat), content_type='application/json')



@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def arpit(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = json.loads( request.body.decode('utf-8') )
	money = Money.objects.get(pk=1)
	print(data)
	if data['data']['type'] == "subtract":
		money.twothousand -= int(data['data']['deno_2000'])
		money.fivehundred -= int(data['data']['deno_500'])
		money.twohundred -= int(data['data']['deno_200'])
		money.hundred -= int(data['data']['deno_100'])
		money.fifty -= int(data['data']['deno_50'])
	if data['data']['type'] == "update":
		money.twothousand = int(data['data']['deno_2000'])
		money.fivehundred = int(data['data']['deno_500'])
		money.twohundred = int(data['data']['deno_200'])
		money.hundred = int(data['data']['deno_100'])
		money.fifty = int(data['data']['deno_50'])
	money.save()
	dat = {"success":1}
	return HttpResponse(json.dumps(dat), content_type='application/json')
	


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def check_updates(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data=[]
	for gr in Group.objects.all():
		b=[]
		pl = Enteredplayer.objects.filter(group=gr).filter(controls_passed=False).filter(controls_displayed=False)
		a=[]
		for p in pl:
			a.append(Regplayer.objects.get(pk=p.regplayer_id))
			p.controls_displayed = True
			p.save()
		for t in a:
			s=[]
			s.append(t.city)
			s.append(t.name.name)
			s.append(t.email_id)
			s.append(t.gender)
			s.append(t.unbilled_amt)
			s.append(t.college)
			s.append(t.mobile_no)
			s.append(t.entered)
			s.append(t.sport)
			s.append(t.pk)
			b.append(s)
		if b:
			data.append({"participants":b,"groupid":gr.group_code})
	return HttpResponse(json.dumps(data), content_type='application/json')


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def piyali(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		print(data['data']['id_arr'])
		for i in data['data']['id_arr']:
			rp = Regplayer.objects.get(pk=int(i))
			pl = Enteredplayer.objects.get(regplayer = rp)
			pl.controls_passed = True
			pl.billcontrols = Billcontrols.objects.filter(unbilled_amt=0).first()
			pl.save()
		bil = Billcontrols.objects.filter(unbilled_amt=0).first()
		bil.unbilled_amt = data['data']['paid_amt']
		bil.amt_received = data['data']['paid_amt']
		bil.dd_no = data['data']['dd_num']
		bil.save()
		dat = {"success":1}
		print(pl.controls_passed)

		data_recnacc=[]
		for gr in Group.objects.all():
			b=[]
			pl = Enteredplayer.objects.filter(controls_passed=True).filter(group=gr).filter(recnacc_displayed=False)
			a=[]
			for p in pl:
				a.append(Regplayer.objects.get(pk=p.regplayer_id))
				p.recnacc_displayed = True
				p.save()
			for t in a:
				b.append({"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
			if b:
				data_recnacc.append({"participants":b,"groupid":gr.group_code})
				
		print("pusher start")
		pusher_client.trigger('my-channel', 'my-event', data_recnacc)
		print("pusher")
		return HttpResponse(json.dumps(dat), content_type='application/json')


@cache_page(CACHE_TTL)
@login_required(login_url = '/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def unconfirm_grp(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'controls/controls-dereg/index.html')


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def unconfirm_details(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data=[]
	for gr in Group.objects.all():
		b=[]
		pl = Enteredplayer.objects.filter(group=gr).filter(controls_passed=True).filter(recnacc_passed=False)
		a=[]
		for p in pl:
			a.append(Regplayer.objects.get(pk=p.regplayer_id))
			p.controls_displayed = True
			p.save()
		for t in a:
			s=[]
			s.append(t.city)
			s.append(t.name.name)
			s.append(t.email_id)
			s.append(t.gender)
			s.append(t.unbilled_amt)
			s.append(t.college)
			s.append(t.mobile_no)
			s.append(t.entered)
			s.append(t.sport)
			s.append(t.pk)
			b.append(s)
		if b:
			data.append({"participants":b,"groupid":gr.group_code})
	return HttpResponse(json.dumps(data), content_type='application/json')


#var data = [["BITS Pilani",[["Arpit", 9829775537],["Nikhil", 921379131],["Sri", 2349719891],["Satya", 9958295537]]],["BITS Hyderabad",[["Arpit", 9829775537],["Nikhil", 921379131],["Sri", 2349719891],["Satya", 9958295537],["Piyali", 4567890435]]],["IIT Delhi",[["Part1", 47656575537],["Part2", 7647676],["Part3", 2345678435678]]]];
@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def view_stats(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for t in Team.objects.all():
		dat = []
		das = []
		for us in User.objects.filter(team=t,deleted=0):
			try:	
				rp = Regplayer.objects.get(name=us)
				pl = Enteredplayer.objects.get(regplayer=rp)
				if pl.controls_passed is True:
					d = []
					d.append(rp.name.name)
					d.append(rp.mobile_no)
					das.append(d)
			except:
				pass
		if das:
			dat.append(t.college)
			dat.append(das)
		if dat:
			data.append(dat)
	print(data)
	return HttpResponse(json.dumps(data), content_type='application/json')


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def unconfirm_player(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		for i in data['data']['id_arr']:
			rp = Regplayer.objects.get(pk=int(i))
			pl = Enteredplayer.objects.get(regplayer = rp)
			pl.controls_passed = False
			pl.billcontrols = None
			pl.save()
		dat = {"success":1}
		return HttpResponse(json.dumps(dat), content_type='application/json')


@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def passed_stats(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	fire_conf = Enteredplayer.objects.all().count()
	cont_conf = Enteredplayer.objects.filter(controls_passed=True).count()
	rec_conf = Enteredplayer.objects.filter(recnacc_passed=True).count()
	data = {"fire_conf":fire_conf,"cont_conf":cont_conf,"rec_conf":rec_conf}
	return HttpResponse(json.dumps(data), content_type='application/json')
	
@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def stats_excel(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=Controls_stats.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "Controls Passed Stats"

	row_num = 0

	columns = [
		(u"ID", 15),
		(u"Name", 40),
		(u"College",50),
		(u"Phone", 20),
		(u"Email", 50),
		(u"Sport", 20),
	]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		#c.style.font.bold = True
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for obj in Enteredplayer.objects.filter(controls_passed=True):
		row_num += 1
		row = [
			obj.regplayer.pk,
			obj.regplayer.name.name,
			obj.regplayer.college,
			obj.regplayer.mobile_no,
			obj.regplayer.email_id,
			obj.regplayer.sport,
		]

		for col_num in range(len(row)):
				c = ws.cell(row=row_num + 1, column=col_num + 1)
				c.value = row[col_num]

	wb.save(response)
	return response

@login_required(login_url='/regsoft/')
@user_passes_test(is_controls_admin, login_url='/regsoft/')
def stats_csv(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	
	response = HttpResponse(content_type='text/csv')
	#decide the file name
	response['Content-Disposition'] = 'attachment; filename="Controls_stats.csv"'

	writer = csv.writer(response, csv.excel)
	response.write(u'\ufeff'.encode('utf8'))

	writer.writerow([
		smart_str(u"ID"),
		smart_str(u"Name"),
		smart_str(u"College"),
		smart_str(u"Phone"),
		smart_str(u"Email"),
		smart_str(u"Sport"),
	])

	for obj in Enteredplayer.objects.filter(controls_passed=True):
		writer.writerow([
			smart_str(obj.regplayer.pk),
			smart_str(obj.regplayer.name.name),
			smart_str(obj.regplayer.college),
			smart_str(obj.regplayer.mobile_no),
			smart_str(obj.regplayer.email_id),
			smart_str(obj.regplayer.sport),
		])
	return response


#def render_to_pdf(request):
#	data = []
#	for obj in Enteredplayer.objects.filter(controls_passed=True):
#		data.append({"pk":obj.regplayer.pk,"name":obj.regplayer.name.name,"college":obj.regplayer.college,"mobile_no":obj.regplayer.mobile_no,"email_id":obj.regplayer.email_id,"sport":obj.regplayer.sport})
#	return render(request,'controls/controls_stats.html',{"mylist":data})

#def html_to_pdf_view(request):
 #   paragraphs = ['first paragraph', 'second paragraph', 'third paragraph']
 #   html_string = render_to_string('controls/controls_stats.html', {'paragraphs': paragraphs})

  #  html = HTML(string=html_string)
   # html.write_pdf(target='/tmp/controls_stats.pdf');

#    fs = FileSystemStorage('/tmp')
#    with fs.open('controls_stats.pdf') as pdf:
#        response = HttpResponse(pdf, content_type='application/pdf')
#        response['Content-Disposition'] = 'attachment; filename="controls_stats.pdf"'
#        return response
 
 #   return response

def stats_html(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for obj in Enteredplayer.objects.filter(controls_passed=True):
		data.append({"pk":obj.regplayer.pk,"name":obj.regplayer.name.name,"college":obj.regplayer.college,"mobile_no":obj.regplayer.mobile_no,"email_id":obj.regplayer.email_id,"sport":obj.regplayer.sport})
	context = {"mylist":data}
	return render(request,'controls/controls_stats.html',context)


def bill_pdf(request,bill_pk):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	print("compres lite")
	data = []
	billl = Billcontrols.objects.get(pk=int(bill_pk))
	for obj in billl.enteredplayer_set.all():
		if obj.regplayer.name.grp_leader == 1:
			data.append({"lis":{"name":obj.regplayer.name.name,"mobile":obj.regplayer.name.phone,"college":obj.regplayer.college}})
		data.append({"pk":obj.regplayer.pk,"name":obj.regplayer.name.name,"unbilled_amt":obj.regplayer.unbilled_amt})
	context = {"grand_total":billl.unbilled_amt,"amt_paid":billl.amt_received,"amt_returned":(billl.amt_received-billl.unbilled_amt),"mylist":data,"time":str(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}
	return render(request,'controls/bill_pdf.html',context)


def denomination_display(request):
	if request.user.is_authenticated():
		if is_controls_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	money = Money.objects.get(pk=1)
	data = []
	data.apped(money.fifty)
	data.append(money.hundred)
	data.append(money.twohundred)
	data.append(money.fivehundred)
	data.append(money.twothousand)
	return HttpResponse(json.dumps(data), content_type='application/json')


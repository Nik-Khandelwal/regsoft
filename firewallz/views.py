from __future__ import unicode_literals
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from main.models import Group,Regplayer,Enteredplayer,Sport,Team
from django.contrib.auth import logout
from django.contrib.auth.models import User
import json
from django.contrib.auth import get_user_model
import string
import random
from django.core import serializers
import pusher
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.shortcuts import render
from django.views.decorators.cache import cache_page
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.decorators import login_required, user_passes_test

from random import choice
from string import ascii_uppercase
import re
import openpyxl
from openpyxl.utils import get_column_letter
import csv
from django.utils.encoding import smart_str
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404,HttpResponseRedirect, JsonResponse
from main.models import CustomUser, Team
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.views.generic import View, ListView, FormView
from django.views import generic
from register.forms import LoginForm, TeamForm, PlayerForm
from django.db import IntegrityError
from django.contrib.auth.models import User
import json
from django.core import serializers, mail
from django.contrib.auth.decorators import login_required
from django.dispatch import receiver
from django.db.models.functions import Lower
from random import choice
from string import ascii_uppercase

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from register.tokens import account_activation_token
from django.core.mail import EmailMessage
from django.core import serializers
from main.models import Group,Regplayer,Enteredplayer,Sport,Firewallz_user

from django.contrib.auth import get_user_model
User=get_user_model()
 

#CACHE_TTL = getattr(settings, '#CACHE_TTL', DEFAULT_TIMEOUT)


pusher_client = pusher.Pusher(
  app_id='499153',
  key='9b825df805e0b694cccc',
  secret='f2bbd60c69e36c90a572',
  cluster='ap2',
  ssl=True
)


def replaceindex(text,index=0,replacement=''):
    return '%s%s%s'%(text[:index],replacement,text[index+1:])


def is_firewallz_admin(user):
	if user:
		if Firewallz_user.objects.get(pk=1).user == user:
			return True
	return False



def group_code_generator(size=5, chars=string.ascii_uppercase + string.digits):
	str = ''.join(random.choice(chars) for _ in range(size))
	for grp in Group.objects.all():
		if grp.group_code == str:
			str = group_code_generator()
	return str


#@cache_page(#CACHE_TTL)
@login_required(login_url = '/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def main(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'firewallz/index.html')


@login_required(login_url = '/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def details(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		dat=[]
		for rp in Regplayer.objects.filter(entered=False):
			# rp.unbilled_amt = 1100-int(rp.name.pcramt)
			b = {"name":rp.name.name,"gender":rp.gender,"college":rp.college,"city":rp.city,"mobile_no":rp.mobile_no,"email_id":rp.email_id,"sport":rp.sport,"entered":rp.entered,"unbilled_amt":rp.unbilled_amt}
			dat.append({"pk":rp.pk,"fields":b})
		#data = serializers.serialize("json", Regplayer.objects.filter(entered=False))
		return HttpResponse(json.dumps(dat), content_type='application/json')

#def confirm_group(request,player_list):
#	grp = Group()
#	grp.group_code = group_code_generator()
#	grp.save()
#	for i in range(1,len(player_list.split("a"))):
#		player_id = int(player_list.split("a")[i])
#		Player = Regplayer.objects.get(pk=player_id)
#		Player.entered = True
#		Player.save()
#		pl = Enteredplayer()
#		pl.regplayer = Player
#		pl.group = grp
#		pl.save()
#	data = {"groupcode":grp.group_code}
#	return HttpResponse(json.dumps(data), content_type='application/json')


@login_required(login_url = '/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def confirm_group(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	grp = Group()
	grp.group_code = group_code_generator()
	grp.save()
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		print(data)
		for dt in data['data']:	
			print(dt['pk'])
			Player = Regplayer.objects.get(pk=dt['pk'])
			print(Player.name.name)
			Player.entered = True
			Player.unbilled_amt = 1100-int(Player.name.pcramt)
			Player.name.grp_leader = dt['groupleader']
			Player.save()
			pl = Enteredplayer()
			pl.regplayer = Player
			pl.group = grp
			pl.save()
			if dt['groupleader'] == 1:
				grp.group_leader = int(dt['pk'])
				grp.save()
		data = {"groupcode":grp.group_code,"pk":grp.pk}
		return HttpResponse(json.dumps(data), content_type='application/json')


def confirm_group_pusher(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data_controls=[]
	for gr in Group.objects.all():
		b=[]
		pl = Enteredplayer.objects.filter(group=gr).filter(controls_passed=False).filter(controls_displayed=False)
		a=[]
		for p in pl:
			a.append(Regplayer.objects.get(pk=p.regplayer_id))
			p.controls_displayed = True
			p.save()
		for t in a:
			s=[]
			s.append(t.city)
			s.append(t.name.name)
			s.append(t.email_id)
			s.append(t.gender)
			s.append(t.unbilled_amt)
			s.append(t.college)
			s.append(t.mobile_no)
			s.append(t.entered)
			s.append(t.sport)
			s.append(t.pk)
			b.append(s)
		if b:
			data_controls.append({"participants":b,"groupid":gr.group_code})
	print("confirm_group_pusher started")
	pusher_client.trigger('my-channel2', 'my-event2', data_controls)
	return HttpResponse(json.dumps({"success":1}), content_type='application/json')




@login_required(login_url = '/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def add_participant(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = 	json.loads( request.body.decode('utf-8') )
		up = User()
		up.save()
		up.username= (''.join(choice(ascii_uppercase) for i in range(5))) + str(up.pk)
		passworduser=(''.join(choice(ascii_uppercase) for i in range(12)))+str(up.pk)
		up.save()
		up.set_password(passworduser)
		up.name = data['data'][0]['name']
		up.email = data['data'][0]['email']
		up.confirm1 = 2
		up.gender = data['data'][0]['gender']
		up.save()
		pl = Regplayer()
		pl.name = up
		pl.gender = data['data'][0]['gender']
		tm = Team.objects.get(pk=int(data['data'][0]['college']))
		pl.college = tm.college
		pl.city = tm.city
		pl.mobile_no = str(data['data'][0]['phone'])
		pl.email_id = data['data'][0]['email']
		#pl.bitsian = data['data'][0]['bitsian']
		pl.entered = False
		pl.sport=''
		for i in data['data'][0]['sport']: 
			up.sportid=replaceindex(up.sportid,int(i),'2')
			sp=Sport.objects.get(idno=int(i))
			up.sport.add(sp)
			up.save()
			pl.sport=pl.sport+Sport.objects.get(idno=int(i)).sport+','
		pl.save()
		pl.uid = "18CB"+str(100000+pl.pk)[-4:]
		pl.save()

		if(pl.bitsian == True):
			pl.unbilled_amt = 0
		pl.save()
		
		to_email = up.email
		message = render_to_string('register/msg2.html', {
										'user':up.name, 
										'username':up.username,
										'password':passworduser,
										
										})
		mail_subject = 'Your account details.'
		email = EmailMessage(mail_subject, message, to=[to_email])
		#email.send()
		dat = {"pk":pl.pk,"college":pl.college}
		return HttpResponse(json.dumps(dat), content_type='application/json')


#@cache_page(#CACHE_TTL)
@login_required(login_url = '/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def unconfirm_grp(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'firewallz/firewallz_unconfirm/index.html')


def unconfirm_details(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for gr in Group.objects.all():
		if gr.enteredplayer_set.filter(controls_passed=False).count() is not 0:
			if Regplayer.objects.get(pk=gr.group_leader):
				player = Regplayer.objects.get(pk=gr.group_leader)
				data.append({"pk":gr.pk,"name": player.name.name, "college":player.college, "groupid":gr.group_code})
	return HttpResponse(json.dumps(data), content_type='application/json')


def show_details_unconfirm(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = 	json.loads( request.body.decode('utf-8') )
		print(data)
		dat = []
		gr = Group.objects.get(group_code=data['data']['group_id'])
		for pl in Enteredplayer.objects.filter(group=gr).filter(controls_passed=False):
			dat.append({"name":pl.regplayer.name.name,"college":pl.regplayer.college, "id":pl.regplayer.pk})
		return HttpResponse(json.dumps(dat), content_type='application/json')


def unconfirm_player(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		print(data)
		dat = []
		pl = Regplayer.objects.get(pk=data['data']['participant_id'])
		pl.entered = False
		pl.save()
		en = Enteredplayer.objects.get(regplayer=pl)
		en.delete()
		datss = []
		b = {"name":pl.name.name,"gender":pl.gender,"college":pl.college,"city":pl.city,"mobile_no":pl.mobile_no,"email_id":pl.email_id,"sport":pl.sport,"entered":pl.entered,"unbilled_amt":pl.unbilled_amt}
		datss.append({"pk":pl.pk,"fields":b})
		print("unconfirm_player pusher")
		print(datss)
		pusher_client.trigger('firewallz_unconfirm_channel', 'firewallz_unconfirm_event', datss)
		dat = {"success":1}
		return HttpResponse(json.dumps(dat), content_type='application/json')

def unconfirm_player_grp(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = json.loads( request.body.decode('utf-8') )
		print(data)
		datss = []
		gr = Group.objects.get(pk=data['data']['group_id'])
		for pl in gr.enteredplayer_set.filter(controls_passed=False):
			rp = Regplayer.objects.get(pk=pl.regplayer.pk)
			rp.entered = False
			rp.save()
			pl.delete()
			b = {"name":rp.name.name,"gender":rp.gender,"college":rp.college,"city":rp.city,"mobile_no":rp.mobile_no,"email_id":rp.email_id,"sport":rp.sport,"entered":rp.entered,"unbilled_amt":rp.unbilled_amt}
			datss.append({"pk":rp.pk,"fields":b})
		print("unconfirm_player_grp pusher")
		print(datss)
		pusher_client.trigger('firewallz_unconfirm_channel', 'firewallz_unconfirm_event', datss)
		dat = {"success":1}
		return HttpResponse(json.dumps(dat), content_type='application/json')



def sportlist(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = []
		for sp in Sport.objects.all():
			data.append({"pk":sp.pk,"sport":sp.sport})
		return HttpResponse(json.dumps(data), content_type='application/json')	


@login_required(login_url='/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def passed_stats(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	fire_conf = Enteredplayer.objects.all().count()
	cont_conf = Enteredplayer.objects.filter(controls_passed=True).count()
	rec_conf = Enteredplayer.objects.filter(recnacc_passed=True).count()
	data = {"fire_conf":fire_conf,"cont_conf":cont_conf,"rec_conf":rec_conf}
	return HttpResponse(json.dumps(data), content_type='application/json')


def id_card(request,string):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	dats = []
	print(string)
	for pl in Enteredplayer.objects.filter(group = Group.objects.get(group_code=string)):
		dats.append({"uid":pl.regplayer.uid,"pk":pl.regplayer.pk,"name":pl.regplayer.name.name,"college":pl.regplayer.college,"group_code":string,"group_id":pl.group_id,"sport":pl.regplayer.sport,"mobile_no":pl.regplayer.mobile_no})
	return render(request,'firewallz/id_template.html',{"data":dats})


def collegelist(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for tm in Team.objects.order_by(Lower('college')):
		data.append({"pk":tm.pk,"college":tm.college})
	return JsonResponse({"data":data})

 #var data = [["BITS Pilani",[["Arpit", 9829775537],["Nikhil", 921379131],["Sri", 2349719891],["Satya", 9958295537]]],["BITS Hyderabad",[["Arpit", 9829775537],["Nikhil", 921379131],["Sri", 2349719891],["Satya", 9958295537],["Piyali", 4567890435]]],["IIT Delhi",[["Part1", 47656575537],["Part2", 7647676],["Part3", 2345678435678]]]];


def view_stats(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data =[]
	for tm in Team.objects.filter(activate=1).order_by(Lower('college')):
		dat = []
		da = []
		for pl in Regplayer.objects.filter(college = tm.college):
			d = []
			try:
				en = Enteredplayer.objects.get(regplayer=pl)
			except Enteredplayer.DoesNotExist:
				en = None
			if en != None:
				d.append(pl.name.name)
				d.append(pl.mobile_no)
			if d:
				da.append(d)
		if da:
			dat.append(tm.college)
			dat.append(da)
			data.append(dat)
	return JsonResponse({"data":data})

def stats_excel(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=Firewallz_stats.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "Firewallz Passed Stats"

	row_num = 0

	columns = [
		(u"ID", 15),
		(u"Name", 40),
		(u"Group_code", 20),
		(u"College",50),
		(u"Phone", 20),
		(u"Email", 50),
		(u"Sport", 20),
	]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		#c.style.font.bold = True
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for obj in Enteredplayer.objects.all():
		row_num += 1
		row = [
			obj.regplayer.pk,
			obj.regplayer.name.name,
			obj.group.group_code,
			obj.regplayer.college,
			obj.regplayer.mobile_no,
			obj.regplayer.email_id,
			obj.regplayer.sport,
		]

		for col_num in range(len(row)):
				c = ws.cell(row=row_num + 1, column=col_num + 1)
				c.value = row[col_num]

	wb.save(response)
	return response


@login_required(login_url='/regsoft/')
@user_passes_test(is_firewallz_admin, login_url='/regsoft/')
def stats_csv(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	
	response = HttpResponse(content_type='text/csv')
	#decide the file name
	response['Content-Disposition'] = 'attachment; filename="Firewallz_stats.csv"'

	writer = csv.writer(response, csv.excel)
	response.write(u'\ufeff'.encode('utf8'))

	writer.writerow([
		smart_str(u"ID"),
		smart_str(u"Name"),
		smart_str(u"Group_Code"),
		smart_str(u"College"),
		smart_str(u"Phone"),
		smart_str(u"Email"),
		smart_str(u"Sport"),
	])

	for obj in Enteredplayer.objects.all():
		writer.writerow([
			smart_str(obj.regplayer.pk),
			smart_str(obj.regplayer.name.name),
			smart_str(obj.group.group_code),
			smart_str(obj.regplayer.college),
			smart_str(obj.regplayer.mobile_no),
			smart_str(obj.regplayer.email_id),
			smart_str(obj.regplayer.sport),
		])
	return response


def stats_html(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for obj in Enteredplayer.objects.all():
		data.append({"pk":obj.regplayer.pk,"name":obj.regplayer.name.name,"group_code":obj.group.group_code,"college":obj.regplayer.college,"mobile_no":obj.regplayer.mobile_no,"email_id":obj.regplayer.email_id,"sport":obj.regplayer.sport})
	context = {"mylist":data}
	return render(request,'firewallz/firewallz_stats.html',context)

def view_id_card(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'firewallz/id_index.html')

def view_id_card_details(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for gr in Group.objects.all():
		if gr.enteredplayer_set.count() is not 0:
			if Regplayer.objects.get(pk=gr.group_leader):
				player = Regplayer.objects.get(pk=gr.group_leader)
				data.append({"pk":gr.pk,"name": player.name.name, "college":player.college, "groupid":gr.group_code})
	return HttpResponse(json.dumps(data), content_type='application/json')

def view_id_card_show_details(request):
	if request.user.is_authenticated():
		if is_firewallz_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	if request.method=='POST':
		data = 	json.loads( request.body.decode('utf-8') )
		print(data)
		dat = []
		gr = Group.objects.get(group_code=data['data']['group_id'])
		for pl in Enteredplayer.objects.filter(group=gr):
			dat.append({"name":pl.regplayer.name.name,"college":pl.regplayer.college, "id":pl.regplayer.pk})
		return HttpResponse(json.dumps(dat), content_type='application/json')
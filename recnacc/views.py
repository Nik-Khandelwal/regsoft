from __future__ import unicode_literals
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from main.models import Note, Group,Regplayer,Enteredplayer,Billcontrols,Accorecnacc,Accomodation,Controlsystem,Singleroom,Acco_name
from django.contrib.auth import logout
from django.contrib.auth.models import User
import json
from django.contrib.auth import get_user_model
import string
import random
from django.core import serializers
import pusher
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.shortcuts import render
from django.views.decorators.cache import cache_page
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from main.models import Group,Regplayer,Enteredplayer,Sport,Team
from django.contrib.auth import logout
from django.contrib.auth.models import User
import json
from django.contrib.auth import get_user_model
import string
import random
from django.core import serializers
import pusher
from django.conf import settings
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.shortcuts import render
from django.views.decorators.cache import cache_page
from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse
from django.template.loader import render_to_string
from io import StringIO
from xhtml2pdf import pisa
from django.template.loader import get_template
from django.template import Context
from django.http import HttpResponse
from cgi import escape
#from weasyprint import HTML
from random import choice
from string import ascii_uppercase
import re
import openpyxl
from openpyxl.utils import get_column_letter
import csv
from django.utils.encoding import smart_str
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, Http404,HttpResponseRedirect, JsonResponse
from main.models import CustomUser, Team
from django.contrib.auth import authenticate, login, logout
from django.contrib import auth
from django.views.generic import View, ListView, FormView
from django.views import generic
from register.forms import LoginForm, TeamForm, PlayerForm
from django.db import IntegrityError
from django.contrib.auth.models import User
import json
from django.core import serializers, mail
from django.contrib.auth.decorators import login_required
from django.dispatch import receiver

from random import choice
from string import ascii_uppercase

from django.contrib.sites.shortcuts import get_current_site
from django.utils.encoding import force_bytes, force_text
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from register.tokens import account_activation_token
from django.core.mail import EmailMessage
from django.core import serializers
from main.models import Group,Regplayer,Enteredplayer,Sport,Money,Billcontrols,Recnacc_user
from django.contrib.auth import logout

from django.contrib.auth import get_user_model
User=get_user_model()

CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)

pusher_client = pusher.Pusher(
  app_id='499153',
  key='9b825df805e0b694cccc',
  secret='f2bbd60c69e36c90a572',
  cluster='ap2',
  ssl=True
)

def is_recnacc_admin(user):
	if user:
		if Recnacc_user.objects.get(pk=1).user == user:
			return True
	return False


@cache_page(CACHE_TTL)
@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def main(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("main")
			cr = Controlsystem()
			cr.save()
			return render(request,'recnacc/index.html')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def participant_details(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("participant_details")
			data=[]
			for gr in Group.objects.all():
				b=[]
				pl = Enteredplayer.objects.filter(controls_passed=True).filter(group=gr).filter(recnacc_passed=False)
				#print(pl)
				a=[]
				for p in pl:
					a.append(Regplayer.objects.get(pk=p.regplayer_id))
					p.recnacc_displayed = True
					p.save()
					#print(p)
				for t in a:
					b.append({"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
				if b:
					data.append({"participants":b,"groupid":gr.group_code})
			#print(data)
			return HttpResponse(json.dumps(data), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')



@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def acco_details(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("acco_details")
			data={}
			b=[]
			for ho in Accomodation.objects.all():
				if 'Single' not in ho.name:
					if ho.vacancy > 0:
						b.append({"id":ho.pk, "name":ho.name, "no":ho.vacancy,"mf":ho.mf})
				else:
					c=[]
					for sr in ho.singleroom.all():
						c.append({"name":sr.name})
					if ho.vacancy > 0:
						b.append({"id":ho.pk, "name":ho.name, "no":ho.vacancy,"rooms":c,"mf":ho.mf})
			data = {"fields":b}
			return HttpResponse(json.dumps(data), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')



@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def srivatsa(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("srivatsa")
			bill = Accorecnacc()
			bill.save()
			dat = {"success":1}
			return HttpResponse(json.dumps(dat), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def satyavrat(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("satyavrat")
			if request.method=='POST':
				data = json.loads( request.body.decode('utf-8') )
				#print(data)
				for i in range(0,int(data['data']['num'])-1):
					bill = Accorecnacc()
					bill.save()
				dat = {"success":1}
				return HttpResponse(json.dumps(dat), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def accomodate(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("accomodate")
			if request.method=='POST':
				data = json.loads( request.body.decode('utf-8') )
				for i in data['data']['id_arr']:
					#print(data['data']['id_arr'])
					rp = Regplayer.objects.get(pk=int(i))
					pl = Enteredplayer.objects.get(regplayer=rp)
					pl.accorecnacc = Accorecnacc.objects.filter(accomodation=None).first()
					pl.recnacc_passed = True
					pl.save()
					dat = {"success":1}
				ac = Accomodation.objects.get(pk=data['data']['bhawan_select'])
				ac.vacancy -= len(data['data']['id_arr'])
				ac.save()
				bill = Accorecnacc.objects.filter(accomodation=None).first()
				bill.accomodation = ac
				bill.save()
				data_update = [4]
				pusher_client.trigger('recnacc_channel', 'recnacc_event', data_update)
				return HttpResponse(json.dumps(dat), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')



@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def accomodate_singleroom(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("accomodate_singleroom")
			if request.method=='POST':
				data = json.loads( request.body.decode('utf-8') )
				for i,j in zip(data['data']['id_arr'], data['data']['bhawan_select']):
					bill = Accorecnacc.objects.filter(accomodation=None).first()
					rp = Regplayer.objects.get(pk=int(i))
					pl = Enteredplayer.objects.get(regplayer=rp)
					sr = Singleroom.objects.get(name=j)
					sr.vacancy += 1
					sr.save()
					ac = sr.accomodation_set.all()[0]
					ac.vacancy -= 1
					ac.save()
					bill.singleroom = sr
					bill.accomodation = ac
					bill.save()
					pl.accorecnacc = bill
					pl.recnacc_passed = True
					pl.save()
				dat = {"success":1}
				data_update = [5]
				pusher_client.trigger('recnacc_channel', 'recnacc_event', data_update)
				return HttpResponse(json.dumps(dat), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')



@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def check_updates(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("check_updates")
			data=[]
			for gr in Group.objects.all():
				b=[]
				pl = Enteredplayer.objects.filter(controls_passed=True).filter(group=gr).filter(recnacc_displayed=False)
				a=[]
				for p in pl:
					a.append(Regplayer.objects.get(pk=p.regplayer_id))
					p.recnacc_displayed = True
					p.save()
				for t in a:
					b.append({"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
				if b:
					data.append({"participants":b,"groupid":gr.group_code})
			return HttpResponse(json.dumps(data), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@cache_page(CACHE_TTL)
@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def unconfirm_acco(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("unconfirm_acco")
			return render(request,'recnacc/Recn_De_Acc/index.html')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def unconfirm_acco_details(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("unconfirm_acco_details")
			data=[]
			for gr in Group.objects.all():
				b=[]
				pl = Enteredplayer.objects.filter(controls_passed=True).filter(group=gr).filter(recnacc_passed=True).filter(all_done=False)
				#print(pl)
				a=[]
				for p in pl:
					t = Regplayer.objects.get(pk=p.regplayer_id)
					hos = str(p.accorecnacc.accomodation)
					if p.accorecnacc.singleroom:
						hos+=str(p.accorecnacc.singleroom)
					b.append({"hostel":hos,"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
					p.recnacc_displayed = True
					p.save()
					#print(p)
				# for t in a:
				# 	b.append({"hostel":,"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
				if b:
					data.append({"participants":b,"groupid":gr.group_code})
				#print(data)
			return HttpResponse(json.dumps(data), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def deaccomodate(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("deaccomodate")
			if request.method=='POST':
				data = json.loads( request.body.decode('utf-8') )
				print(data)
				fne = 0
				dats = []
				for i in data['data']['id_arr']:
					#print(data['data']['id_arr'])
					rp = Regplayer.objects.get(pk=int(i))
					pl = Enteredplayer.objects.get(regplayer=rp)
					pl.all_done = True
					pl.save()
					dats.append({"name":rp.name.name,"fine":rp.fine})
					fne += rp.fine
					if pl.accorecnacc.accomodation is not None:
						ac = Accomodation.objects.get(pk=pl.accorecnacc.accomodation.pk)
						ac.vacancy += 1
						ac.save()
					elif pl.accorecnacc.singleroom is not None:
						ac = Accomodation.objects.get(pk=pl.accorecnacc.singleroom.pk)
						ac.vacancy += 1
						ac.save()
					else:
						pass

				dat = {"total":fne,"list":dats}
				data_update = [7]
				pusher_client.trigger('recndeacc_channel', 'recndeacc_event', data_update)
				return HttpResponse(json.dumps(dat), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def redeaccomodate(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = json.loads( request.body.decode('utf-8') )
	rp = Regplayer.objects.get(pk=int(data['data']['pk']))
	pl = Enteredplayer.objects.get(regplayer=rp)
	pl.all_done = False
	pl.save()
	if pl.accorecnacc.accomodation is not None:
		ac = Accomodation.objects.get(pk=pl.accorecnacc.accomodation.pk)
		ac.vacancy -= 1
		ac.save()
	elif pl.accorecnacc.singleroom is not None:
		ac = Accomodation.objects.get(pk=pl.accorecnacc.singleroom.pk)
		ac.vacancy -= 1
		ac.save()
	else:
		pass
	data_update = [4]
	pusher_client.trigger('recnacc_channel', 'recnacc_event', data_update)
	return HttpResponse(json.dumps({"success":1}), content_type='application/json')




@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def fine_amount(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("fine_amount")
			if request.method=='POST':
				data = json.loads( request.body.decode('utf-8') )
				for i in data['data']['id_arr']:
					#print(data['data']['id_arr'])
					rp = Regplayer.objects.get(pk=int(i))
					rp.fine += data['data']['fine_amount']
					rp.save()
				dat = {"success":1}
				return HttpResponse(json.dumps(dat), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


#def deaccomodate(request):
#	data = json.loads( request.body.decode('utf-8') )
#	for i in data['data']['id_arr']:
#		#print(data['data']['id_arr'])
#		rp = Regplayer.objects.get(pk=int(i))
#		pl = Enteredplayer.objects.get(regplayer=rp)
#		ac = pl.accorecnacc
#		if ac.accomodation is not None:
#			ac.accomodation.vacancy += 1
#			ac.save()
#			pl.accorecnacc = None
#			pl.save()
#		if ac.singleroom is not None:
#			ac.singleroom.vacancy += 1
#			ac.save()
#		pl.recnacc_passed = False
#		pl.save()
#		dat = {"success":1}
#	return HttpResponse(json.dumps(dat), content_type='application/json')

@cache_page(CACHE_TTL)
@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def reconfirm_acco(request):
	print("reconfirm_acco1")
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("reconfirm_acco")
			return render(request,'recnacc/Recn_Re_Acc/index.html')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def reconfirm_acco_details(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("reconfirm_acco_details")
			data=[]
			for gr in Group.objects.all():
				b=[]
				pl = Enteredplayer.objects.filter(controls_passed=True).filter(group=gr).filter(recnacc_passed=True).filter(all_done=False)
				#print(pl)
				a=[]
				for p in pl:
					t = Regplayer.objects.get(pk=p.regplayer_id)
					hos = str(p.accorecnacc.accomodation)
					if p.accorecnacc.singleroom:
						hos+=str(p.accorecnacc.singleroom)
					b.append({"hostel":hos,"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
					p.recnacc_displayed = True
					p.save()
					#print(p)
				# for t in a:
				# 	b.append({"hostel":,"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
				if b:
					data.append({"participants":b,"groupid":gr.group_code})
			return HttpResponse(json.dumps(data), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def reaccomodate(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("reaccomodate")
			if request.method=='POST':
				data = json.loads( request.body.decode('utf-8') )
				fine = 0
				for i in data['data']['id_arr']:
					#print(data['data']['id_arr'])
					rp = Regplayer.objects.get(pk=int(i))
					pl = Enteredplayer.objects.get(regplayer=rp)
					if pl.accorecnacc.accomodation is not None:
						pl.accorecnacc.accomodation.vacancy += 1
					elif pl.accorecnacc.singleroom is not None:
						pl.accorecnacc.singleroom.vacancy += 1
					else:
						pass

					pl.accorecnacc = None
					pl.recnacc_passed = False
					pl.save()
				return HttpResponse(json.dumps({"success":1}), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


def reaccomodate_pusher(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("participant_details")
			dat=[]
			for gr in Group.objects.all():
				b=[]
				a=[]
				data = json.loads( request.body.decode('utf-8') )
				for i in data['data']:
					#print(data['data']['id_arr'])
					rp = Regplayer.objects.get(pk=int(i))
					pl = Enteredplayer.objects.get(regplayer=rp)
					if pl.group == gr:
						a.append(rp)
						pl.recnacc_displayed = True
						pl.save()
					#print(p)
				for t in a:
					b.append({"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
				if b:
					dat.append({"participants":b,"groupid":gr.group_code})
			#print(data)
			pusher_client.trigger('recnreacc_channel', 'recnreacc_event', dat)

		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def passed_stats(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("passed_stats")
			fire_conf = Enteredplayer.objects.all().count()
			cont_conf = Enteredplayer.objects.filter(controls_passed=True).count()
			rec_conf = Enteredplayer.objects.filter(recnacc_passed=True).count()
			data = {"fire_conf":fire_conf,"cont_conf":cont_conf,"rec_conf":rec_conf}
			return HttpResponse(json.dumps(data), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
 

#var jsonResponse = {"data": [["Ram", ["Common Room", 90, "Single Rooms", 40, "TT Room", 80]],["Budh", ["Common Room", 120, "Single Rooms", 50]],["Meera", ["Common Room", 190, "Single Rooms", 100]],["MAL-A", ["Common Room", 90, "Single Rooms", 20, "TT Room", 10]],["Ram", ["Common Room", 90, "Single Rooms", 40, "TT Room", 80]]]};
@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def availability_stats(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			print("availability_stats")
			data = []
			for ac in Acco_name.objects.all():
				das = []
				das.append(ac.name)
				dat = []
				dat.append("Common Room")
				if(ac.common_room):
					dat.append(ac.common_room.vacancy)
				else:
					dat.append(0)
				dat.append("Single Rooms")
				if(ac.s_room):
					dat.append(ac.s_room.vacancy)
				else:
					dat.append(0)
				dat.append("TT Room")
				if(ac.tt_room):
					dat.append(ac.tt_room.vacancy)
				else:
					dat.append(0)
				das.append(dat)
				data.append(das)

			dt = {"data":data}
			return HttpResponse(json.dumps(dt), content_type='application/json')
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')

#var jsonResponse = {"data": [["Ram", [["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]]]],["Ram", [["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]]]],["Ram", [["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]]]],["Ram", [["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]],["Common Room", ["Arpit Anshuman", 1, "Nikhil Khandelwal", 2, "Srivatsa", 3]]]]]};


# def view_stats(request):
# 	print("view_stats")
# 	data = []
# 	for ac in Acco_name.objects.all():
# 		das = []
# 		das.append(ac.name)
# 		dat = []
# 		dt = []
# 		for pl in Enteredplayer.objects.all():
# 			try:
# 				pt = pl.accorecnacc
# 				if pt.accomodation == ac.common_room:
# 					dt.append(pl.regplayer.name.name)
# 					dt.append(pl.regplayer.pk)
# 					dat.append(dt)
# 					print(dat)
# 			except:
# 				pass		
# 		das.append(dat)
# 		data.append

# 		return HttpResponse(json.dumps({"data":data}), content_type='application/json')


def view_stats(request):
	print("view_stats")
	data = []
	for ac in Acco_name.objects.all():
		dat = []
		for pl in Enteredplayer.objects.all():
			dic = {}
			try:
				pt = pl.accorecnacc
				if pt.accomodation == ac.common_room:
					dic = {"type":"common_room","name":pl.regplayer.name.name,"mobile":pl.regplayer.mobile_no}
				elif pt.accomodation == ac.tt_room:
					dic = {"type":"tt_room","name":pl.regplayer.name.name,"mobile":pl.regplayer.mobile_no}
				elif pt.accomodation == ac.s_room:
					dic = {"type":"s_room","room_no":pt.singleroom.name,"name":pl.regplayer.name.name,"mobile":pl.regplayer.mobile_no}
				else:
					pass
				dat.append(dic)
			except:
				pass	
		data.append({"hostel_name":ac.name,"list":dat})
	return HttpResponse(json.dumps({"data":data}), content_type='application/json')


# Bhawan Occupency

@cache_page(CACHE_TTL)
def acco_strength(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'recnacc/occupancy.html')


def disp_occupency(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			dats = []
			for ac in Acco_name.objects.all():
				dat = []
				dat.append(ac.name)
				if(ac.common_room):
					dat.append({"pk":ac.common_room.pk,"strength":ac.common_room.strength,"fine":ac.common_room.fine})
				else:
					dat.append({"pk":0,"strength":0,"fine":0})
				if(ac.tt_room):
					dat.append({"pk":ac.tt_room.pk,"strength":ac.tt_room.strength,"fine":ac.tt_room.fine})
				else:
					dat.append({"pk":0,"strength":0,"fine":0})
				if(ac.s_room):
					dat.append({"pk":ac.s_room.pk,"strength":ac.s_room.strength,"fine":ac.s_room.fine})
				else:
					dat.append({"pk":0,"strength":0,"fine":0})
				dat.append(ac.pk)
				dats.append(dat)
			return HttpResponse(json.dumps({"data":dats}), content_type='application/json')	
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


def edit_occupency(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			data = json.loads( request.body.decode('utf-8') )
			print(data)
			yo = int(data['data']['strength'])
			ac = Accomodation.objects.get(pk=data['data']['pk'])
			diff = yo - ac.strength
			ac.strength = yo

			diff = int(data['data']['strength']) - ac.strength
			ac.strength = data['data']['strength']

			ac.vacancy += diff
			ac.save()
			data_update = [9]
			pusher_client.trigger('recnacc_occupancy_channel', 'recnacc_occupancy_event', data_update)
			return HttpResponse(json.dumps({"success":"1"}), content_type='application/json')	
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')

@cache_page(CACHE_TTL)
def deallocated(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request, 'recnacc/deallocated.html')

		
def deallocated_page(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	# data = []
	# for pl in Enteredplayer.objects.filter(all_done=True):
	# 	data.append({"name":pl.regplayer.name.name,"college":pl.regplayer.college})
	# return HttpResponse(json.dumps({"data":data}), content_type='application/json')	
	data=[]
	for gr in Group.objects.all():
		b=[]
		pl = gr.enteredplayer_set.filter(all_done=True)
		#print(pl)
		a=[]
		for p in pl:
			t = Regplayer.objects.get(pk=p.regplayer.pk)
			hos = str(p.accorecnacc.accomodation)
			if p.accorecnacc.singleroom:
				hos+=str(p.accorecnacc.singleroom)
			b.append({"hostel":hos,"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
			p.recnacc_displayed = True
			p.save()
			#print(p)
		#for t in a:
		# 	b.append({"hostel":,"indiv_name":t.name.name, "indiv_college":t.college, "indiv_gender":t.gender, "indiv_id":t.pk})
		if b:
			data.append({"participants":b,"groupid":gr.group_code})
	return HttpResponse(json.dumps(data), content_type='application/json')
		

@cache_page(CACHE_TTL)
def fines(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'recnacc/fines.html')

def fine_page(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			data = json.loads( request.body.decode('utf-8') )
			ac = Accomodation.objects.get(pk=int(data['data']['pk']))
			ac.fine += float(data['data']['amt'])
			ac.save()
			cnt = 0
			for ar in ac.accorecnacc_set.all():
				for pl in Enteredplayer.objects.filter(all_done=False):
					if pl.accorecnacc == ar:
						cnt+=1

			for ar in ac.accorecnacc_set.all():
				for pl in Enteredplayer.objects.filter(all_done=False):
					if pl.accorecnacc == ar:
						rp = Regplayer.objects.get(pk=pl.regplayer.pk)
						rp.fine += (float(data['data']['amt'])/cnt)
						rp.save()

			return HttpResponse(json.dumps({"success":"1"}), content_type='application/json')	
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')

@cache_page(CACHE_TTL)
def notes(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	return render(request,'recnacc/notes.html')

def view_notes(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			data = []
			for n in Note.objects.all():
				data.append({"pk":n.pk,"time":n.time.strftime('%d-%m-%Y %H:%M:%S UTC'),"text":n.text})
			return HttpResponse(json.dumps({"data":data}), content_type='application/json')	
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')

def add_note(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			data = json.loads( request.body.decode('utf-8') )
			n = Note()
			n.text = data['data']['text']
			n.save()
			return HttpResponse(json.dumps({"success":"1"}), content_type='application/json')	
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')


def stats_excel(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
	response['Content-Disposition'] = 'attachment; filename=Recnacc_stats.xlsx'
	wb = openpyxl.Workbook()
	ws = wb.get_active_sheet()
	ws.title = "Recnacc Passed Stats"

	row_num = 0

	columns = [
		(u"ID", 15),
		(u"Hostel", 10),
		(u"Room", 20),
		(u"Name",50),
		(u"College", 50),
		(u"Mobile_no", 50),
	]

	for col_num in range(len(columns)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = columns[col_num][0]
		#c.style.font.bold = True
		# set column width
		ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

	for ac in Acco_name.objects.all():
		for pl in Enteredplayer.objects.all():
			row_num += 1
			try:
				pt = pl.accorecnacc
				if pt.accomodation == ac.common_room:
					row = [
						pl.regplayer.pk,
						ac.name,
						"common_room",
						pl.regplayer.name.name,
						pl.regplayer.college,
						pl.regplayer.mobile_no,
						]
				elif pt.accomodation == ac.tt_room:
					row = [
						pl.regplayer.pk,
						ac.name,
						"tt_room",
						pl.regplayer.name.name,
						pl.regplayer.college,
						pl.regplayer.mobile_no,
						]
				elif pt.accomodation == ac.s_room:
					row = [
						pl.regplayer.pk,
						ac.name,
						pt.singleroom.name,
						pl.regplayer.name.name,
						pl.regplayer.college,
						pl.regplayer.mobile_no,
						]
				else:
					pass
			except:
				pass

	for col_num in range(len(row)):
		c = ws.cell(row=row_num + 1, column=col_num + 1)
		c.value = row[col_num]

	wb.save(response)
	return response


@login_required(login_url='/regsoft/')
@user_passes_test(is_recnacc_admin, login_url='/regsoft/')
def stats_csv(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	
	response = HttpResponse(content_type='text/csv')
	#decide the file name
	response['Content-Disposition'] = 'attachment; filename="Recnacc_stats.csv"'

	writer = csv.writer(response, csv.excel)
	response.write(u'\ufeff'.encode('utf8'))

	row_num =0

	writer.writerow([
		smart_str(u"ID"),
		smart_str(u"Hostel"),
		smart_str(u"Room"),
		smart_str(u"Name"),
		smart_str(u"College"),
		smart_str(u"Mobile_no"),
	])

	for ac in Acco_name.objects.all():
		for pl in Enteredplayer.objects.all():
			row_num += 1
			try:
				pt = pl.accorecnacc
				if pt.accomodation == ac.common_room:
					writer.writerow([
						smart_str(pl.regplayer.pk),
						smart_str(ac.name),
						smart_str("common_room"),
						smart_str(pl.regplayer.name.name),
						smart_str(pl.regplayer.college),
						smart_str(pl.regplayer.mobile_no),
						])
				elif pt.accomodation == ac.tt_room:
					writer.writerow([
						smart_str(pl.regplayer.pk),
						smart_str(ac.name),
						smart_str("tt_room"),
						smart_str(pl.regplayer.name.name),
						smart_str(pl.regplayer.college),
						smart_str(pl.regplayer.mobile_no),
						])
				elif pt.accomodation == ac.s_room:
					writer.writerow([
						smart_str(pl.regplayer.pk),
						smart_str(ac.name),
						smart_str(pt.singleroom.name),
						smart_str(pl.regplayer.name.name),
						smart_str(pl.regplayer.college),
						smart_str(pl.regplayer.mobile_no),
						])
				else:
					pass
			except:
				pass
	return response


def stats_html(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = []
	for obj in Enteredplayer.objects.filter(recnacc_passed=True):
		data.append({"pk":obj.regplayer.pk,"name":obj.regplayer.name.name,"group_code":obj.group.group_code,"college":obj.regplayer.college,"mobile_no":obj.regplayer.mobile_no,"email_id":obj.regplayer.email_id,"sport":obj.regplayer.sport})
	context = {"mylist":data}
	return render(request,'recnacc/recnacc_stats.html',context)


def delete_note(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = json.loads( request.body.decode('utf-8') )
	n = Note.objects.get(pk=data['data']['pk'])
	n.delete()
	return HttpResponse(json.dumps({"success":"1"}), content_type='application/json')	


def add_bhawan(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	data = json.loads( request.body.decode('utf-8') )
	acn = Acco_name()
	acn.name = data['data']['hostel']
	acn.save()
	return HttpResponse(json.dumps({"success":"1"}), content_type='application/json')


def add_acco(request):
	if request.user.is_authenticated():
		if is_recnacc_admin(request.user):
			pass
		else:
			logout(request)
			return HttpResponseRedirect('/regsoft/')
	else:
		return HttpResponseRedirect('/regsoft/')
	success = 1
	data = json.loads( request.body.decode('utf-8') )
	acn = Acco_name.objects.get(pk=data['data']['pk'])
	if(int(data['data']['type']) == 0):
		if(acn.common_room):
			success = 0
		else:
			ac = Accomodation()
			ac.name = data['data']['ac_name']
			ac.strength = data['data']['ac_strength']
			ac.mf = data['data']['gender']
			ac.vacancy = data['data']['ac_strength']
			ac.save()
			acn.common_room = ac
			acn.save()
	elif(int(data['data']['type']) == 1):
		if(ac.tt_room):
			success = 0
		else:
			ac = Accomodation()
			ac.name = data['data']['ac_name']
			ac.strength = data['data']['ac_strength']
			ac.mf = data['data']['gender']
			ac.vacancy = data['data']['ac_strength']
			ac.save()
			acn.tt_room = ac
			acn.save()
	elif(int(data['data']['type']) == 2):
		if(acn.s_room):
			ac = acn.s_room
			ac.strength += data['data']['sr_vacancy']
			ac.vacancy += data['data']['sr_vacancy']
			ac.save()
		else:
			ac = Accomodation()
			ac.name = data['data']['ac_name']
			ac.mf = data['data']['gender']
			ac.strength = data['data']['sr_vacancy']
			ac.vacancy = data['data']['sr_vacancy']
			ac.save()
			acn.s_room = ac
			acn.save()
		sr = Singleroom()
		sr.name = data['data']['sr_name']
		sr.vacancy = data['data']['sr_vacancy']
		sr.save()
		ac.singleroom.add(sr)
		ac.save()
		acn.s_room = ac
	acn.save()
	return HttpResponse(json.dumps({"success":success}), content_type='application/json')